import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from datetime import datetime, date
import io

from common import CYCLE_STATUSES, WORKFLOW_STATUSES, duration_to_hours, get_current_and_previous_sprints


# === EXCEL FORMATTER ===
def format_excel(df, output_file_label, cycle_threshold, lead_threshold):
    if cycle_threshold <= 0 or lead_threshold <= 0:
        raise ValueError("Cycle Time and Lead Time thresholds must be positive integers.")
    
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="JIRA Cycle Times")
    
    wb = load_workbook(output_buffer)
    ws = wb.active
    ws.title = "JIRA Cycle Times"
    ws.sheet_properties.tabColor = "1072BA"

    format_sheet(ws, df.columns, cycle_threshold, lead_threshold, output_file_label)
    
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# === FORMAT SHEET ===
def format_sheet(sheet, headers, cycle_threshold, lead_threshold, output_file_label):
    create_table(sheet)
    freeze_top_row(sheet)
    auto_adjust_column_width(sheet)
    align_headers(sheet)
    selected_team_name_for_sprints = output_file_label.split('_')[0].replace('-', ' ').title()
    highlight_current_sprint_multiline(sheet, headers, selected_team_name_for_sprints) 
    highlight_long_durations(sheet, cycle_threshold, lead_threshold)

# === TABLE CREATION AND FORMATTING ===
def create_table(sheet):
    table = Table(displayName="JIRAMetricsTable", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
    style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

# === SHEET FORMATTING FUNCTIONS ===
def freeze_top_row(sheet):
    sheet.freeze_panes = "B2"

def auto_adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 5
    set_border(sheet, f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
    sheet.sheet_view.showGridLines = False

def align_headers(sheet):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(color="FFFFFF", bold=True)

def set_border(sheet, cell_range):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = thin_border

def highlight_long_durations(sheet, cycle_threshold, lead_threshold):
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    col_idx = map_columns(sheet)
    add_tooltip_comments(sheet, col_idx, cycle_threshold, lead_threshold)
    apply_story_points_gradient(sheet, col_idx)
    highlight_rows(sheet, col_idx, orange_fill, cycle_threshold, lead_threshold)
    add_legend(sheet, orange_fill, cycle_threshold, lead_threshold)

def map_columns(sheet):
    header = [cell.value for cell in sheet[1]]
    return {col: idx + 1 for idx, col in enumerate(header)}

def add_tooltip_comments(sheet, col_idx, cycle_threshold, lead_threshold):
    if col_idx.get("Cycle Time"):
        sheet.cell(row=1, column=col_idx["Cycle Time"]).comment = Comment(f"Orange if Cycle Time > {cycle_threshold // 24} days", "System")
    if col_idx.get("Lead Time"):
        sheet.cell(row=1, column=col_idx["Lead Time"]).comment = Comment(f"Orange if Lead Time > {lead_threshold // 24} days", "System")
    if col_idx.get("Story Points"):
        sheet.cell(row=1, column=col_idx["Story Points"]).comment = Comment("Green gradient: low → high Story Points", "System")

def apply_story_points_gradient(sheet, col_idx):
    sp_col = col_idx.get("Story Points")
    if sp_col:
        sp_letter = get_column_letter(sp_col)
        sp_range = f"{sp_letter}2:{sp_letter}{sheet.max_row}"
        blue_gradient = ColorScaleRule(
            start_type='min', start_color='E6F0FA',
            mid_type='percentile', mid_value=50, mid_color='4FC3F7',
            end_type='max', end_color='1565C0'
        )
        sheet.conditional_formatting.add(sp_range, blue_gradient)

def get_column_index_by_header(sheet, header_name):
    for cell in sheet[1]:
        if str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return -1

def highlight_current_sprint_multiline(sheet, headers, team_name_for_sprint, log_list):
    current_sprint_full, previous_sprint_full = get_current_and_previous_sprints(team_name_for_sprint, base_sprint="2025.12", base_start_date_str="2025-06-11")

    col_idx = get_column_index_by_header(sheet, "Sprints")
    if col_idx == -1:
        log_list.append(f"WARN Sprints column not found for sprint highlighting.")
        return
    
    for row in sheet.iter_rows(min_row=2):
        cell = row[col_idx - 1]

        if not cell.value:
            continue

        sprint_values_in_cell = [s.strip() for s in str(cell.value).split(",")]
        modified = False
        updated_sprints = []

        for sprint_text in sprint_values_in_cell:
            sprint_clean = sprint_text.replace("🔶", "").replace("🔷", "").strip()

            if sprint_clean == current_sprint_full:
                updated_sprints.append(f"{sprint_clean} 🔶")
                modified = True
            elif sprint_clean == previous_sprint_full:
                updated_sprints.append(f"{sprint_clean} 🔷")
                modified = True
            else:
                updated_sprints.append(sprint_text)

        if modified:
            cell.value = ", ".join(updated_sprints)

def get_duration_hours_from_excel_cell(sheet, row, col):
    if col:
        val = sheet.cell(row=row, column=col).value
        return duration_to_hours(val) if isinstance(val, str) and val.strip().upper() == "N/A" else None
    return None

def highlight_cell(sheet, row, col, hours, threshold, fill):
    if col and hours is not None and hours > threshold:
        sheet.cell(row=row, column=col).fill = fill

def is_threshold_breached(hours, threshold):
    return hours is not None and hours >= threshold

def calculate_cycle_time_hours_from_excel(sheet, row, col_idx):
    total_hours = 0
    for status in CYCLE_STATUSES:
        col = col_idx.get(status)
        if not col: continue
        val = sheet.cell(row=row, column=col).value
        if isinstance(val, str) and val.strip().upper() != "N/A":
            hours_in_status = duration_to_hours(val)
            if hours_in_status is not None: total_hours += hours_in_status
    return total_hours if total_hours > 0 else None

def highlight_rows(sheet, col_idx, orange_fill, cycle_threshold, lead_threshold):
    cycle_time_header_col_idx = col_idx.get("Cycle Time")
    lead_time_header_col_idx = col_idx.get("Lead Time")

    for row in range(2, sheet.max_row + 1):
        current_cycle_hours = calculate_cycle_time_hours_from_excel(sheet, row, col_idx)
        current_lead_hours = get_duration_hours_from_excel_cell(sheet, row, lead_time_header_col_idx)

        highlight_cell(sheet, row, cycle_time_header_col_idx, current_cycle_hours, cycle_threshold, orange_fill)
        highlight_cell(sheet, row, lead_time_header_col_idx, current_lead_hours, lead_threshold, orange_fill)

        if should_apply_heatmap(current_cycle_hours, cycle_threshold, current_lead_hours, lead_threshold):
            breach_scope = determine_breach_scope(current_cycle_hours, cycle_threshold, current_lead_hours, lead_threshold)
            apply_workflow_heatmap(sheet, row, col_idx, breach_scope)

def should_apply_heatmap(cycle_hours, cycle_threshold, lead_hours, lead_threshold):
    return (is_threshold_breached(cycle_hours, cycle_threshold) or is_threshold_breached(lead_hours, lead_threshold))

def determine_breach_scope(cycle_hours, cycle_threshold, lead_hours, lead_threshold):
    lead_breach = is_threshold_breached(lead_hours, lead_threshold)
    cycle_breach = is_threshold_breached(cycle_hours, cycle_threshold)
    if lead_breach: return "lead"
    elif cycle_breach: return "cycle"
    return None

def apply_workflow_heatmap(sheet, row, col_idx, scope="lead"):
    if scope is None: return
    if scope == "cycle": workflow_subset = CYCLE_STATUSES
    elif scope == "lead": workflow_subset = [status for status in WORKFLOW_STATUSES if status not in {"Released", "Closed"}]
    else: workflow_subset = WORKFLOW_STATUSES

    row_durations = {}; values = []
    for status in workflow_subset:
        col = col_idx.get(status)
        if not col: continue
        val = sheet.cell(row=row, column=col).value
        hours = duration_to_hours(val)
        row_durations[status] = (col, hours)
        if hours is not None: values.append(hours)
    
    if not values: return

    min_val, max_val = min(values), max(values)
    delta = max_val - min_val if max_val != min_val else 1

    for status, (col, hours) in row_durations.items():
        if hours is None: continue
        intensity = (hours - min_val) / delta
        hex_color = calculate_heatmap_color(intensity)
        sheet.cell(row=row, column=col).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def calculate_heatmap_color(intensity):
    r = 255; g = int(200 - 120 * intensity); b = int(200 - 120 * intensity)
    r = max(0, min(255, r)); g = max(0, min(255, g)); b = max(0, min(255, b))
    return f"FF{r:02X}{g:02X}{b:02X}"

def calculate_heatmap_color_blue_gradient(intensity):
    r_start, g_start, b_start = (230, 240, 250)
    r_end, g_end, b_end = (21, 101, 192)

    r = int(r_start + (r_end - r_start) * intensity)
    g = int(g_start + (g_end - g_start) * intensity)
    b = int(b_start + (b_end - b_start) * intensity)
    
    return f"FF{r:02X}{g:02X}{b:02X}"

def add_legend(sheet, orange_fill, cycle_threshold, lead_threshold):
    legend_col = sheet.max_column + 2
    sheet.cell(row=1, column=legend_col, value="Legend").font = Font(bold=True, size=12, underline="single")
    legends = [
        (f"Cycle Time > {cycle_threshold // 24}d", orange_fill),
        (f"Lead Time > {lead_threshold // 24}d", orange_fill),
        ("Story Points: Light→Dark Blue", PatternFill(start_color="A9D0F5", end_color="1565C0", fill_type="solid")),
        ("Workflow: Light→Dark Red (per row, if breached)", PatternFill(start_color="FFCCCC", end_color="FF6666", fill_type="solid")),
    ]
    for i, (label, fill) in enumerate(legends, start=2):
        cell = sheet.cell(row=i, column=legend_col, value=label)
        cell.fill = fill
        cell.font = Font(bold=True)
    adjust_legend_column_width(sheet, legend_col)

def adjust_legend_column_width(sheet, legend_col):
    legend_letter = get_column_letter(legend_col)
    max_len = max((len(str(sheet.cell(row=row, column=legend_col).value)) for row in range(1, sheet.max_row + 1) if sheet.cell(row=row, column=legend_col).value), default=0)
    sheet.column_dimensions[legend_letter].width = max_len + 5
